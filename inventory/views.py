# inventory/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
from django.urls import reverse_lazy, reverse
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.db.models.functions import TruncMonth
from django.contrib.contenttypes.models import ContentType
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.template.loader import render_to_string, get_template
from django.http import HttpResponse, JsonResponse, FileResponse, HttpResponseBadRequest
from django.conf import settings
from django.db import models, connection
from django.contrib.auth.models import User, Group, Permission
from django.views.decorators.http import require_POST
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth import login

from .models import (
    Employee, ITAsset, Department, Position, AssetType, 
    OwnerCompany, AssetHistory, Notification, NotificationCategory,
    Outlet
)
from .forms import (
    EmployeeForm, ITAssetForm, SuperUserRegistrationForm,
    AssetAssignForm, OwnerCompanyForm, UserRegistrationForm
)

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from xhtml2pdf import pisa
from weasyprint import HTML, CSS
from weasyprint.text.fonts import FontConfiguration
from datetime import datetime, timedelta
import tempfile
import json
import io
from io import BytesIO
import os
import zipfile
from django.core import management

@login_required
def home(request):
    # Get total counts
    total_assets = ITAsset.objects.count()
    total_employees = Employee.objects.count()
    available_assets = ITAsset.objects.filter(status='available').count()
    assigned_assets = ITAsset.objects.filter(status='assigned').count()
    maintenance_assets = ITAsset.objects.filter(status='maintenance').count()
    retired_assets = ITAsset.objects.filter(status='retired').count()

    # Get asset type distribution
    asset_types = AssetType.objects.all()
    asset_type_distribution = []
    for asset_type in asset_types:
        count = ITAsset.objects.filter(asset_type=asset_type).count()
        asset_type_distribution.append({
            'name': asset_type.display_name,
            'count': count
        })

    # Get owner company distribution
    owner_companies = OwnerCompany.objects.all()
    owner_company_distribution = []
    for company in owner_companies:
        count = ITAsset.objects.filter(owner=company).count()
        owner_company_distribution.append({
            'name': company.name,
            'count': count
        })

    # Get department distribution
    departments = Department.objects.all()
    department_distribution = []
    for department in departments:
        count = Employee.objects.filter(department=department).count()
        department_distribution.append({
            'name': department.name,
            'count': count
        })

    # Get recent assets
    recent_assets = ITAsset.objects.select_related('asset_type', 'owner', 'assigned_to').order_by('-id')[:5]

    # Get assets with expiring warranty (within 3 months)
    today = datetime.now().date()
    three_months_later = today + timedelta(days=90)
    expiring_warranty_assets = ITAsset.objects.filter(
        warranty_expiry__isnull=False,
        warranty_expiry__gte=today,
        warranty_expiry__lte=three_months_later
    ).select_related('asset_type', 'owner', 'assigned_to').order_by('warranty_expiry')

    # Get asset history
    asset_history = AssetHistory.objects.select_related(
        'asset', 'asset__asset_type', 'employee'
    ).order_by('-date')[:50]  # Show last 50 history entries

    context = {
        'total_assets': total_assets,
        'total_employees': total_employees,
        'available_assets': available_assets,
        'assigned_assets': assigned_assets,
        'maintenance_assets': maintenance_assets,
        'retired_assets': retired_assets,
        'asset_type_distribution': asset_type_distribution,
        'owner_company_distribution': owner_company_distribution,
        'department_distribution': department_distribution,
        'recent_assets': recent_assets,
        'expiring_warranty_assets': expiring_warranty_assets,
        'asset_history': asset_history,
    }
    return render(request, 'inventory/home.html', context)



@login_required
def employee_create(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            employee = form.save()
            messages.success(request, f'Employee "{employee.get_full_name()}" was successfully created.')
            return redirect('inventory:employee_detail', pk=employee.pk)
    else:
        form = EmployeeForm()
    
    return render(request, 'employees/employee_form.html', {'form': form})

@login_required
def employee_detail(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    return render(request, 'employees/employee_detail.html', {'employee': employee})

@login_required
def employee_update(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            employee = form.save()
            messages.success(request, f'Employee "{employee.get_full_name()}" was successfully updated.')
            return redirect('inventory:employee_detail', pk=employee.pk)
    else:
        form = EmployeeForm(instance=employee)
    
    return render(request, 'employees/employee_form.html', {'form': form, 'employee': employee})

@login_required
def employee_delete(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    
    if request.method == 'POST':
        employee_name = employee.get_full_name()
        employee.delete()
        messages.success(request, f'Employee "{employee_name}" was successfully deleted.')
        return redirect('inventory:employee_list')
    
    return render(request, 'employees/employee_confirm_delete.html', {'employee': employee})

@login_required
def employee_toggle_status(request, pk):
    """Toggle employee's active status."""
    employee = get_object_or_404(Employee, pk=pk)
    
    if request.method == 'POST':
        employee.is_active = not employee.is_active
        employee.save()
        
        status = 'activated' if employee.is_active else 'deactivated'
        messages.success(request, f'Employee "{employee.get_full_name()}" has been {status}.')
    
    return redirect('inventory:employee_detail', pk=employee.pk)

@login_required
def asset_assign(request):
    """View for assigning assets to employees."""
    if request.method == 'POST':
        asset_id = request.POST.get('asset')
        employee_id = request.POST.get('employee')
        receipt_date = request.POST.get('receipt_date')
        
        try:
            asset = ITAsset.objects.get(pk=asset_id)
            employee = Employee.objects.get(pk=employee_id)
            
            # Create history record before updating asset
            AssetHistory.objects.create(
                asset=asset,
                action='assigned',
                employee=employee,
                notes=f'Asset assigned to {employee.get_full_name()}',
                created_by=request.user
            )
            
            # Get or create Device Assignment category
            device_category, _ = NotificationCategory.objects.get_or_create(
                name='Device Assignment',
                defaults={
                    'icon': 'fa-laptop',
                    'color': 'primary',
                    'description': 'Notifications about device assignments and returns'
                }
            )
            
            # Create notification for the employee
            Notification.objects.create(
                title=f'New Device Assigned: {asset.name}',
                message=(
                    f'You have been assigned a new device:\n'
                    f'Device: {asset.name}\n'
                    f'Serial Number: {asset.serial_number}\n'
                    f'Model: {asset.model}\n'
                    f'Manufacturer: {asset.manufacturer}\n'
                    f'Asset Type: {asset.asset_type.display_name}'
                ),
                category=device_category,
                priority='medium',
                employee_profile=employee,
                content_type=ContentType.objects.get_for_model(asset),
                object_id=asset.id,
                action_url=reverse('inventory:asset_detail', kwargs={'pk': asset.id})
            )
            
            # Update asset status, assignment and receipt date
            asset.assigned_to = employee
            asset.status = 'assigned'
            asset.receipt_date = receipt_date
            asset.save()
            
            messages.success(request, f'Asset "{asset.name}" has been assigned to {employee.get_full_name()}.')
            return redirect('inventory:employee_detail', pk=employee.pk)
            
        except ITAsset.DoesNotExist:
            messages.error(request, 'Invalid asset selected.')
        except Employee.DoesNotExist:
            messages.error(request, 'Invalid employee selected.')
        except Exception as e:
            messages.error(request, f'Error assigning asset: {str(e)}')
        return redirect('inventory:employee_list')
    
    # Get the employee from the query parameter
    employee_id = request.GET.get('employee')
    if employee_id:
        try:
            employee = Employee.objects.get(pk=employee_id)
        except Employee.DoesNotExist:
            messages.error(request, 'Employee not found.')
            return redirect('inventory:employee_list')
    else:
        employee = None
    
    # Get available assets
    available_assets = ITAsset.objects.filter(status='available')
    
    context = {
        'employee': employee,
        'available_assets': available_assets,
    }
    
    return render(request, 'assets/asset_assign.html', context)

class ITAssetListView(LoginRequiredMixin, ListView):
    model = ITAsset
    template_name = 'assets/asset_list.html'
    context_object_name = 'assets'
    paginate_by = 10  # Show 10 items per page

    def get_queryset(self):
        queryset = ITAsset.objects.select_related('asset_type', 'owner', 'assigned_to').all()
        
        # Apply filters from request
        search = self.request.GET.get('search', '')
        asset_type = self.request.GET.get('asset_type', '')
        manufacturer = self.request.GET.get('manufacturer', '')
        status = self.request.GET.get('status', '')
        owner = self.request.GET.get('owner', '')
        
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(serial_number__icontains=search) |
                Q(model__icontains=search) |
                Q(manufacturer__icontains=search)
            )
        if asset_type:
            queryset = queryset.filter(asset_type_id=asset_type)
        if manufacturer:
            queryset = queryset.filter(manufacturer__iexact=manufacturer)
        if status:
            queryset = queryset.filter(status=status)
        if owner:
            queryset = queryset.filter(owner_id=owner)

        return queryset.order_by('-id')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add filter choices to context
        context['asset_types'] = AssetType.objects.all()
        
        # Get unique manufacturers, normalized to title case
        manufacturers = ITAsset.objects.exclude(manufacturer='').values_list('manufacturer', flat=True)
        normalized_manufacturers = set()
        for m in manufacturers:
            if m:
                # Normalize manufacturer name: strip whitespace, convert to title case
                normalized = m.strip().title()
                # Handle common variations
                normalized = normalized.replace('Hp', 'HP')
                normalized = normalized.replace('Ibm', 'IBM')
                normalized = normalized.replace('Dell Inc', 'Dell')
                normalized = normalized.replace('Dell Inc.', 'Dell')
                normalized = normalized.replace('Dell Computer', 'Dell')
                normalized = normalized.replace('Dell Corporation', 'Dell')
                normalized = normalized.replace('Lenovo Group', 'Lenovo')
                normalized = normalized.replace('Lenovo Inc', 'Lenovo')
                normalized = normalized.replace('Lenovo Inc.', 'Lenovo')
                normalized = normalized.replace('Microsoft Corporation', 'Microsoft')
                normalized = normalized.replace('Apple Inc', 'Apple')
                normalized = normalized.replace('Apple Inc.', 'Apple')
                normalized = normalized.replace('Apple Computer', 'Apple')
                normalized_manufacturers.add(normalized)
        
        context['manufacturers'] = sorted(normalized_manufacturers)
        context['status_choices'] = ITAsset.STATUS_CHOICES
        context['owners'] = OwnerCompany.objects.all()
        
        # Add current filter values to context
        context['current_filters'] = {
            'search': self.request.GET.get('search', ''),
            'asset_type': self.request.GET.get('asset_type', ''),
            'manufacturer': self.request.GET.get('manufacturer', ''),
            'status': self.request.GET.get('status', ''),
            'owner': self.request.GET.get('owner', '')
        }
        
        return context

    def get(self, request, *args, **kwargs):
        if request.GET.get('export') == 'excel':
            # Get the filtered queryset
            queryset = self.get_queryset()
            
            # Generate dynamic filename based on filters
            filename_parts = ['assets']
            
            # Add asset type to filename
            asset_type_id = request.GET.get('asset_type')
            if asset_type_id:
                try:
                    asset_type = AssetType.objects.get(id=asset_type_id)
                    filename_parts.append(asset_type.name.lower())
                except AssetType.DoesNotExist:
                    pass
            
            # Add manufacturer to filename
            manufacturer = request.GET.get('manufacturer')
            if manufacturer:
                filename_parts.append(manufacturer.lower())
            
            # Add status to filename
            status = request.GET.get('status')
            if status:
                filename_parts.append(status)
            
            # Add owner to filename
            owner_id = request.GET.get('owner')
            if owner_id:
                try:
                    owner = OwnerCompany.objects.get(id=owner_id)
                    filename_parts.append(owner.code.lower())
                except OwnerCompany.DoesNotExist:
                    pass
            
            # Add timestamp
            filename_parts.append(timezone.now().strftime('%Y%m%d'))
            
            # Create workbook and select active sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Filtered Assets"
            
            # Write headers
            headers = [
                'Asset Name',
                'Asset Type',
                'Serial Number',
                'Model',
                'Manufacturer',
                'Owner Company',
                'Status',
                'Assigned To',
                'Department',
                'Purchase Date',
                'Warranty Expiry'
            ]
            ws.append(headers)
            
            # Write data
            for asset in queryset:
                row = [
                    asset.name,
                    asset.asset_type.display_name if asset.asset_type else '',
                    asset.serial_number,
                    asset.model,
                    asset.manufacturer,
                    asset.owner.name if asset.owner else '',
                    asset.get_status_display(),
                    asset.assigned_to.get_full_name() if asset.assigned_to else '',
                    asset.assigned_to.department.name if asset.assigned_to and asset.assigned_to.department else '',
                    asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '',
                    asset.warranty_expiry.strftime('%Y-%m-%d') if asset.warranty_expiry else ''
                ]
                ws.append(row)
            
            # Create the response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Join filename parts with underscores and add extension
            filename = '_'.join(filename_parts) + '.xlsx'
            response['Content-Disposition'] = f'attachment; filename={filename}'
            
            # Save the workbook to the response
            wb.save(response)
            return response
            
        return super().get(request, *args, **kwargs)

class AssetListContentView(ITAssetListView):
    template_name = 'assets/asset_list_content.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['owners'] = OwnerCompany.objects.all()
        
        if context['is_paginated']:
            paginator = context['paginator']
            current_page = context['page_obj'].number
            total_pages = paginator.num_pages

            # Calculate the range of pages to show (show 5 pages around current page)
            start_page = max(current_page - 2, 1)
            end_page = min(current_page + 2, total_pages)

            # Adjust the range if we're near the start or end
            if start_page <= 3:
                end_page = min(5, total_pages)
            if end_page >= total_pages - 2:
                start_page = max(total_pages - 4, 1)

            context['page_range'] = range(start_page, end_page + 1)
            context['current_page'] = current_page
            context['total_pages'] = total_pages

        return context

    def get(self, request, *args, **kwargs):
        # Handle Excel export through the parent view
        if request.GET.get('export') == 'excel':
            return super().get(request, *args, **kwargs)
            
        # Only serve HTMX requests for the content template
        if 'HX-Request' in request.headers:
            return super().get(request, *args, **kwargs)
            
        return HttpResponseBadRequest('This view only serves HTMX requests')

class ITAssetDetailView(LoginRequiredMixin, DetailView):
    model = ITAsset
    template_name = 'assets/asset_detail.html'
    context_object_name = 'asset'

class ITAssetCreateView(LoginRequiredMixin, CreateView):
    model = ITAsset
    form_class = ITAssetForm
    template_name = 'assets/asset_form.html'
    success_url = reverse_lazy('inventory:asset_list')

    def form_valid(self, form):
        asset = form.save()
        
        try:
            # Create history record for new asset
            AssetHistory.objects.create(
                asset=asset,
                action='received',
                notes=f'Asset received and added to inventory',
                created_by=self.request.user
            )

            # Create notifications if asset is assigned to an employee
            if asset.assigned_to:
                # Get or create Device Assignment category
                device_category, _ = NotificationCategory.objects.get_or_create(
                    name='Device Assignment',
                    defaults={
                        'icon': 'fa-laptop',
                        'color': 'primary',
                        'description': 'Notifications about device assignments and returns'
                    }
                )

                # Create assignment notification
                Notification.objects.create(
                    title=f'New Device Assigned: {asset.name}',
                    message=f'You have been assigned a new device:\nDevice: {asset.name}\nSerial Number: {asset.serial_number}\nModel: {asset.model}',
                    category=device_category,
                    priority='medium',
                    employee_profile=asset.assigned_to,
                    content_type=ContentType.objects.get_for_model(asset),
                    object_id=asset.id,
                    action_url=reverse('inventory:asset_detail', kwargs={'pk': asset.id})
                )

            # Check warranty and create notification if needed
            if asset.warranty_expiry:
                days_remaining = (asset.warranty_expiry - timezone.now().date()).days
                if days_remaining <= 90:  # If warranty expires within 90 days
                    warranty_category, _ = NotificationCategory.objects.get_or_create(
                        name='Warranty',
                        defaults={
                            'icon': 'fa-shield-alt',
                            'color': 'warning',
                            'description': 'Notifications about warranty status'
                        }
                    )

                    priority = 'high' if days_remaining <= 30 else 'medium'
                    message = (
                        f'Warranty Alert for {asset.name}\n'
                        f'Serial Number: {asset.serial_number}\n'
                        f'Days Remaining: {days_remaining}\n'
                        f'Expiry Date: {asset.warranty_expiry}'
                    )

                    Notification.objects.create(
                        title=f'Warranty Expiring Soon: {asset.name}',
                        message=message,
                        category=warranty_category,
                        priority=priority,
                        content_type=ContentType.objects.get_for_model(asset),
                        object_id=asset.id,
                        action_url=reverse('inventory:asset_detail', kwargs={'pk': asset.id})
                    )

            messages.success(self.request, 'IT Asset created successfully.')
        except Exception as e:
            messages.warning(self.request, f'Asset created but there was an error with notifications: {str(e)}')

        return super().form_valid(form)

class ITAssetUpdateView(LoginRequiredMixin, UpdateView):
    model = ITAsset
    form_class = ITAssetForm
    template_name = 'assets/asset_form.html'
    success_url = reverse_lazy('inventory:asset_list')

    def form_valid(self, form):
        # Store old values before saving
        old_status = self.get_object().status
        old_assigned_to = self.get_object().assigned_to
        old_warranty_expiry = self.get_object().warranty_expiry
        old_asset = self.get_object()
        
        response = super().form_valid(form)
        asset = self.object
        
        try:
            # Get or create notification categories
            device_category, _ = NotificationCategory.objects.get_or_create(
                name='Device Assignment',
                defaults={
                    'icon': 'fa-laptop',
                    'color': 'primary',
                    'description': 'Notifications about device assignments and returns'
                }
            )
            
            maintenance_category, _ = NotificationCategory.objects.get_or_create(
                name='Maintenance',
                defaults={
                    'icon': 'fa-tools',
                    'color': 'warning',
                    'description': 'Notifications about device maintenance and repairs'
                }
            )
            
            modification_category, _ = NotificationCategory.objects.get_or_create(
                name='Device Modification',
                defaults={
                    'icon': 'fa-edit',
                    'color': 'info',
                    'description': 'Notifications about device modifications and updates'
                }
            )

            system_category, _ = NotificationCategory.objects.get_or_create(
                name='System',
                defaults={
                    'icon': 'fa-cog',
                    'color': 'secondary',
                    'description': 'System notifications and updates'
                }
            )

            # Handle assignment changes
            if old_assigned_to != asset.assigned_to:
                # If device was unassigned from someone
                if old_assigned_to and not asset.assigned_to:
                    # Notify the previous owner
                    Notification.objects.create(
                        title=f'Device Unassigned: {asset.name}',
                        message=(
                            f'A device has been unassigned from you:\n'
                            f'Device: {asset.name}\n'
                            f'Serial Number: {asset.serial_number}\n'
                            f'Model: {asset.model}\n'
                            f'Manufacturer: {asset.manufacturer}\n\n'
                            f'Please ensure you have returned all accessories and peripherals.'
                        ),
                        category=device_category,
                        priority='medium',
                        employee_profile=old_assigned_to,
                        content_type=ContentType.objects.get_for_model(asset),
                        object_id=asset.id,
                        action_url=reverse('inventory:asset_detail', kwargs={'pk': asset.id})
                    )

                    # Create system notification for unassignment
                    Notification.objects.create(
                        title=f'Device Unassigned: {asset.name}',
                        message=(
                            f'Device has been unassigned:\n'
                            f'Device: {asset.name}\n'
                            f'Serial Number: {asset.serial_number}\n'
                            f'Previous Owner: {old_assigned_to.get_full_name()}\n'
                            f'Action Required: Verify device return and condition.'
                        ),
                        category=device_category,
                        priority='medium',
                        content_type=ContentType.objects.get_for_model(asset),
                        object_id=asset.id,
                        action_url=reverse('inventory:asset_detail', kwargs={'pk': asset.id})
                    )

                # If device was assigned to someone new
                elif asset.assigned_to:
                    # Notify the new owner
                    Notification.objects.create(
                        title=f'New Device Assigned: {asset.name}',
                        message=(
                            f'A new device has been assigned to you:\n'
                            f'Device: {asset.name}\n'
                            f'Serial Number: {asset.serial_number}\n'
                            f'Model: {asset.model}\n'
                            f'Manufacturer: {asset.manufacturer}\n'
                            f'Asset Type: {asset.asset_type.display_name}\n\n'
                            f'Previous Owner: {old_assigned_to.get_full_name() if old_assigned_to else "None"}\n\n'
                            f'Please verify the device and its accessories.\n'
                            f'Contact IT support if you need any assistance setting up your device.'
                        ),
                        category=device_category,
                        priority='high',
                        employee_profile=asset.assigned_to,
                        content_type=ContentType.objects.get_for_model(asset),
                        object_id=asset.id,
                        action_url=reverse('inventory:asset_detail', kwargs={'pk': asset.id})
                    )

                    # Create system notification for new assignment
                    Notification.objects.create(
                        title=f'Device Assigned: {asset.name}',
                        message=(
                            f'Device has been assigned:\n'
                            f'Device: {asset.name}\n'
                            f'Serial Number: {asset.serial_number}\n'
                            f'Assigned To: {asset.assigned_to.get_full_name()}\n'
                            f'Department: {asset.assigned_to.department}\n'
                            f'Previous Owner: {old_assigned_to.get_full_name() if old_assigned_to else "None"}\n'
                            f'Action Required: Schedule device setup and orientation if needed.'
                        ),
                        category=device_category,
                        priority='medium',
                        content_type=ContentType.objects.get_for_model(asset),
                        object_id=asset.id,
                        action_url=reverse('inventory:asset_detail', kwargs={'pk': asset.id})
                    )

                # Create history record for assignment change
                if old_assigned_to:
                    AssetHistory.objects.create(
                        asset=asset,
                        action='unassigned',
                        employee=old_assigned_to,
                        notes=f'Device unassigned from {old_assigned_to.get_full_name()}',
                        created_by=self.request.user
                    )
                if asset.assigned_to:
                    AssetHistory.objects.create(
                        asset=asset,
                        action='assigned',
                        employee=asset.assigned_to,
                        notes=f'Device assigned to {asset.assigned_to.get_full_name()}',
                        created_by=self.request.user
                    )

            # Check for significant modifications
            significant_fields = {
                'name': 'Name',
                'model': 'Model',
                'manufacturer': 'Manufacturer',
                'serial_number': 'Serial Number',
                'mac_address_wifi': 'WiFi MAC Address',
                'mac_address_ethernet': 'Ethernet MAC Address',
                'ip_address': 'IP Address',
                'processor': 'Processor',
                'ram_size': 'RAM Size',
                'hdd1_capacity': 'Primary HDD Capacity',
                'hdd2_capacity': 'Secondary HDD Capacity',
                'operating_system': 'Operating System',
                'delivery_letter_code': 'Delivery Letter Code',
                'purchase_date': 'Purchase Date',
                'receipt_date': 'Receipt Date'
            }
            
            changes = []
            for field, display_name in significant_fields.items():
                old_value = getattr(old_asset, field, None)
                new_value = getattr(asset, field, None)
                if old_value != new_value and (old_value or new_value):  # Check if either value exists
                    changes.append(f"{display_name}: {old_value} → {new_value}")

            # If significant changes were made, create modification notification
            if changes:
                # Notify the assigned employee
                if asset.assigned_to:
                    Notification.objects.create(
                        title=f'Your Device Has Been Modified: {asset.name}',
                        message=(
                            f'Your assigned device has been modified:\n'
                            f'Device: {asset.name}\n'
                            f'Serial Number: {asset.serial_number}\n'
                            f'Changes Made:\n' + '\n'.join(f'• {change}' for change in changes) + '\n\n'
                            f'Please review these changes and report any issues to IT support.'
                        ),
                        category=modification_category,
                        priority='medium',
                        employee_profile=asset.assigned_to,
                        content_type=ContentType.objects.get_for_model(asset),
                        object_id=asset.id,
                        action_url=reverse('inventory:asset_detail', kwargs={'pk': asset.id})
                    )

                # Create system notification for modifications
                Notification.objects.create(
                    title=f'Device Modified: {asset.name}',
                    message=(
                        f'Device has been modified:\n'
                        f'Device: {asset.name}\n'
                        f'Serial Number: {asset.serial_number}\n'
                        f'Modified by: {self.request.user.get_full_name() or self.request.user.username}\n'
                        f'Changes Made:\n' + '\n'.join(f'• {change}' for change in changes)
                    ),
                    category=modification_category,
                    priority='medium',
                    content_type=ContentType.objects.get_for_model(asset),
                    object_id=asset.id,
                    action_url=reverse('inventory:asset_detail', kwargs={'pk': asset.id})
                )

                # Create history record for modification
                AssetHistory.objects.create(
                    asset=asset,
                    action='modified',
                    employee=asset.assigned_to,
                    notes=f'Device modified:\n' + '\n'.join(changes),
                    created_by=self.request.user
                )

            # Handle status changes, maintenance, and retirement
            if old_status != asset.status:
                action = asset.status
                notes = f'Asset status changed from {old_status} to {asset.status}'
                
                # If status changed to maintenance
                if asset.status == 'maintenance':
                    notes = f'Asset sent for maintenance'
                    # Notify the assigned employee
                    if asset.assigned_to:
                        Notification.objects.create(
                            title=f'Your Device Needs Maintenance: {asset.name}',
                            message=(
                                f'Your assigned device has been sent for maintenance:\n'
                                f'Device: {asset.name}\n'
                                f'Serial Number: {asset.serial_number}\n'
                                f'Previous Status: {old_status}\n'
                                f'Expected Duration: To be determined\n'
                                f'Note: A temporary replacement device may be provided if needed.\n\n'
                                f'Please backup any important data if you haven\'t already.'
                            ),
                            category=maintenance_category,
                            priority='high',
                            employee_profile=asset.assigned_to,
                            content_type=ContentType.objects.get_for_model(asset),
                            object_id=asset.id,
                            action_url=reverse('inventory:asset_detail', kwargs={'pk': asset.id})
                        )
                    
                    # Create system notification for maintenance
                    Notification.objects.create(
                        title=f'Device In Maintenance: {asset.name}',
                        message=(
                            f'Device requires maintenance:\n'
                            f'Device: {asset.name}\n'
                            f'Serial Number: {asset.serial_number}\n'
                            f'Previous Status: {old_status}\n'
                            f'Assigned To: {asset.assigned_to.get_full_name() if asset.assigned_to else "Not Assigned"}\n'
                            f'Action Required: Schedule maintenance and update status when complete.'
                        ),
                        category=maintenance_category,
                        priority='high',
                        content_type=ContentType.objects.get_for_model(asset),
                        object_id=asset.id,
                        action_url=reverse('inventory:asset_detail', kwargs={'pk': asset.id})
                    )
                
                # If status changed from maintenance to available/assigned
                elif old_status == 'maintenance' and asset.status in ['available', 'assigned']:
                    notes = f'Asset returned from maintenance'
                    # Notify the assigned employee
                    if asset.assigned_to:
                        Notification.objects.create(
                            title=f'Your Device Has Returned from Maintenance: {asset.name}',
                            message=(
                                f'Your device has completed maintenance:\n'
                                f'Device: {asset.name}\n'
                                f'Serial Number: {asset.serial_number}\n'
                                f'Current Status: {asset.get_status_display()}\n'
                                f'Note: Please verify all your settings and data are intact.'
                            ),
                            category=maintenance_category,
                            priority='medium',
                            employee_profile=asset.assigned_to,
                            content_type=ContentType.objects.get_for_model(asset),
                            object_id=asset.id,
                            action_url=reverse('inventory:asset_detail', kwargs={'pk': asset.id})
                        )

                # If status changed to retired
                elif asset.status == 'retired':
                    notes = f'Asset retired from inventory'
                    
                    # Notify the previously assigned employee
                    if old_assigned_to:
                        Notification.objects.create(
                            title=f'Your Device Has Been Retired: {asset.name}',
                            message=(
                                f'Your previously assigned device has been retired:\n'
                                f'Device: {asset.name}\n'
                                f'Serial Number: {asset.serial_number}\n'
                                f'Model: {asset.model}\n'
                                f'Manufacturer: {asset.manufacturer}\n\n'
                                f'If you need a replacement device, please contact IT support.'
                            ),
                            category=system_category,
                            priority='high',
                            employee_profile=old_assigned_to,
                            content_type=ContentType.objects.get_for_model(asset),
                            object_id=asset.id,
                            action_url=reverse('inventory:asset_detail', kwargs={'pk': asset.id})
                        )

                    # Create system notification for retirement
                    Notification.objects.create(
                        title=f'Device Retired: {asset.name}',
                        message=(
                            f'Device has been retired from inventory:\n'
                            f'Device: {asset.name}\n'
                            f'Serial Number: {asset.serial_number}\n'
                            f'Model: {asset.model}\n'
                            f'Previous Status: {old_status}\n'
                            f'Last Assigned To: {old_assigned_to.get_full_name() if old_assigned_to else "Not Assigned"}\n'
                            f'Action Required: Update inventory records and process device disposal if necessary.'
                        ),
                        category=system_category,
                        priority='medium',
                        content_type=ContentType.objects.get_for_model(asset),
                        object_id=asset.id,
                        action_url=reverse('inventory:asset_detail', kwargs={'pk': asset.id})
                    )
                
                # Create history record for status change
                AssetHistory.objects.create(
                    asset=asset,
                    action=action,
                    employee=asset.assigned_to,
                    notes=notes,
                    created_by=self.request.user
                )

            # Rest of the existing code for warranty notifications
            if asset.warranty_expiry:
                days_remaining = (asset.warranty_expiry - timezone.now().date()).days
                warranty_changed = old_warranty_expiry != asset.warranty_expiry
                
                # Only create notification if:
                # 1. Warranty date has changed OR
                # 2. Warranty is expiring soon (but not already expired)
                if warranty_changed or (0 < days_remaining <= 90):
                    warranty_category, _ = NotificationCategory.objects.get_or_create(
                        name='Warranty',
                        defaults={
                            'icon': 'fa-shield-alt',
                            'color': 'warning',
                            'description': 'Notifications about warranty status'
                        }
                    )

                    # Set appropriate priority and message based on status
                    if days_remaining <= 0:
                        priority = 'high'
                        title = f'Warranty Expired: {asset.name}'
                        message = (
                            f'Warranty has expired for device:\n'
                            f'Device: {asset.name}\n'
                            f'Serial Number: {asset.serial_number}\n'
                            f'Expiry Date: {asset.warranty_expiry}\n'
                            f'Action Required: Please review and update warranty status.'
                        )
                    else:
                        priority = 'high' if days_remaining <= 30 else 'medium'
                        title = f'Warranty Expiring Soon: {asset.name}'
                        message = (
                            f'Warranty Alert for {asset.name}\n'
                            f'Serial Number: {asset.serial_number}\n'
                            f'Days Remaining: {days_remaining}\n'
                            f'Expiry Date: {asset.warranty_expiry}\n'
                            f'Action Required: Please plan for warranty renewal.'
                        )

                    # Only create notification if warranty changed or it's not expired
                    if warranty_changed or days_remaining > 0:
                        Notification.objects.create(
                            title=title,
                            message=message,
                            category=warranty_category,
                            priority=priority,
                            content_type=ContentType.objects.get_for_model(asset),
                            object_id=asset.id,
                            action_url=reverse('inventory:asset_detail', kwargs={'pk': asset.id})
                        )

            messages.success(self.request, 'IT Asset updated successfully.')
        except Exception as e:
            messages.warning(self.request, f'Asset updated but there was an error with notifications: {str(e)}')
        
        return response

class ITAssetDeleteView(LoginRequiredMixin, DeleteView):
    model = ITAsset
    template_name = 'assets/asset_confirm_delete.html'
    success_url = reverse_lazy('inventory:asset_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'IT Asset deleted successfully.')
        return super().delete(request, *args, **kwargs)

@login_required
def employee_upload(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                # Skip header row
                for row in ws.iter_rows(min_row=2):
                    try:
                        # Get or create department first
                        department_name = row[5].value
                        if department_name:
                            department, _ = Department.objects.get_or_create(name=department_name)
                            
                            # Get or create position with department
                            position_name = row[6].value
                            if position_name:
                                position, _ = Position.objects.get_or_create(
                                    name=position_name,
                                    department=department  # Set the department for the position
                                )
                            else:
                                position = None
                        else:
                            department = None
                            position = None
                        
                        # Get company
                        company_name = row[7].value
                        try:
                            company = OwnerCompany.objects.get(name=company_name)
                        except OwnerCompany.DoesNotExist:
                            messages.error(request, f'Row {row[0].row}: Company "{company_name}" not found. Please create it first.')
                            continue
                        
                        # Create employee
                        employee = Employee(
                            employee_id=row[0].value,
                            national_id=row[1].value,
                            first_name=row[2].value,
                            last_name=row[3].value,
                            email=row[4].value,
                            department=department,
                            position=position,
                            company=company,  # Use the company object
                            hire_date=row[8].value
                        )
                        employee.save()
                    except Exception as e:
                        messages.error(request, f'Error processing row: {str(e)}')
                        continue
                
                messages.success(request, 'Employees uploaded successfully!')
                return redirect('inventory:employee_list')
            except Exception as e:
                messages.error(request, f'Error uploading file: {str(e)}')
    
    return render(request, 'employees/employee_upload.html')

@login_required
def download_employee_template(request):
    wb = Workbook()
    ws = wb.active
    
    # Add headers
    headers = [
        'Employee ID *',
        'National ID *',
        'First Name *',
        'Last Name *',
        'Email *',
        'Department *',
        'Position *',
        'Company *',
        'Hire Date *'
    ]
    ws.append(headers)
    
    # Add example row
    example = [
        'EMP001',
        '12345678901234',
        'John',
        'Doe',
        'john.doe@example.com',
        'IT',
        'Software Engineer',
        'CTP',
        '2024-01-01'
    ]
    ws.append(example)
    
    # Add available departments
    departments = Department.objects.all()
    if departments.exists():
        ws.append([])  # Add empty row
        ws.append(['Available Departments:'])
        for dept in departments:
            ws.append([dept.get_name_display()])
    
    # Add available companies
    companies = OwnerCompany.objects.all()
    if companies.exists():
        ws.append([])  # Add empty row
        ws.append(['Available Companies:'])
        for company in companies:
            ws.append([company.name])
    
    # Style the header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Style the sections headers
    if departments.exists():
        dept_header = ws.cell(row=4, column=1)
        dept_header.font = openpyxl.styles.Font(bold=True)
    
    if companies.exists():
        company_header = ws.cell(row=departments.count() + 6, column=1)
        company_header.font = openpyxl.styles.Font(bold=True)
    
    # Add data validation for company column
    company_validation = DataValidation(type="list", formula1=f'"{",".join(companies.values_list("name", flat=True))}"')
    company_validation.add(f'H2:H{len(example) + 1}')
    ws.add_data_validation(company_validation)
    
    # Add data validation for department column
    department_validation = DataValidation(type="list", formula1=f'"{",".join(departments.values_list("name", flat=True))}"')
    department_validation.add(f'F2:F{len(example) + 1}')
    ws.add_data_validation(department_validation)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employee_template.xlsx'
    
    wb.save(response)
    return response

@login_required
def download_employee_data(request):
    wb = Workbook()
    ws = wb.active
    
    # Add headers
    headers = [
        'Employee ID',
        'National ID',
        'First Name',
        'Last Name',
        'Email',
        'Department',
        'Position',
        'Company',
        'Hire Date'
    ]
    ws.append(headers)
    
    # Add employee data
    employees = Employee.objects.all()
    for employee in employees:
        row = [
            employee.employee_id,
            employee.national_id,
            employee.first_name,
            employee.last_name,
            employee.email,
            str(employee.department) if employee.department else '',
            str(employee.position) if employee.position else '',
            str(employee.company.name) if employee.company else '',  # Convert company to string
            employee.hire_date
        ]
        ws.append(row)
    
    # Style the header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=employees.xlsx'
    
    wb.save(response)
    return response

@login_required
def download_asset_template(request):
    wb = Workbook()
    ws = wb.active
    
    # Add headers
    headers = [
        'Asset Name *',
        'Asset Type *',
        'Serial Number *',
        'Model *',
        'Manufacturer',
        'Owner Company *',
        'Assigned Employee ID',
        'Wi-Fi MAC Address',
        'Ethernet MAC Address *',
        'Delivery Letter Code',
        'Purchase Date',
        'Receipt Date',
        'Warranty Expiry',
        'Status *'  # Added status field
    ]
    ws.append(headers)
    
    # Add example row
    example = [
        'Laptop-001',
        'laptop',  # Use the internal name, not display name
        'SN123456',
        'ThinkPad X1',
        'Lenovo',
        'CTP',  # Company name
        'EMP001',  # Employee ID
        '00:1A:2B:3C:4D:5E',
        '00:1A:2B:3C:4D:5F',
        'DL-2024-001',
        '2024-01-01',
        '2024-01-15',
        '2027-01-01',
        'available'  # Default status
    ]
    ws.append(example)
    
    # Add available asset types
    ws.append([])  # Add empty row
    ws.append(['Available Asset Types:'])
    ws.append(['Internal Name - Display Name'])
    for asset_type in AssetType.objects.all().order_by('name'):
        ws.append([f'{asset_type.name} - {asset_type.display_name}'])
    
    # Add available statuses
    ws.append([])  # Add empty row
    ws.append(['Available Statuses:'])
    ws.append(['Internal Name - Display Name'])
    for status_code, status_name in ITAsset.STATUS_CHOICES:
        ws.append([f'{status_code} - {status_name}'])
    
    # Add owner companies section
    ws.append([])  # Add empty row
    ws.append(['Owner Companies:'])
    for company in OwnerCompany.objects.all().order_by('name'):
        ws.append([company.name])
    
    # Add available employees section
    ws.append([])  # Add empty row
    ws.append(['Available Employees:'])
    ws.append(['Format: Employee ID - Full Name (Company)'])
    ws.append([])
    
    # Group employees by company
    for company in OwnerCompany.objects.all():
        ws.append([f'--- {company.name} Employees ---'])
        employees = Employee.objects.filter(
            is_active=True,
            company=company
        ).order_by('employee_id')
        
        for employee in employees:
            ws.append([f"{employee.employee_id} - {employee.get_full_name()} ({company.name})"])
        ws.append([])  # Add empty row between companies
    
    # Style the header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Style the sections headers
    ws.cell(row=4, column=1).font = openpyxl.styles.Font(bold=True)
    ws.cell(row=5, column=1).font = openpyxl.styles.Font(italic=True)  # Style the "Internal Name - Display Name" row
    
    # Find and style other section headers
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        if cell.value in ['Owner Companies:', 'Available Employees:', 'Available Statuses:'] or \
           (cell.value and isinstance(cell.value, str) and cell.value.startswith('---')):
            cell.font = openpyxl.styles.Font(bold=True)
    
    # Add data validation for asset type column
    asset_types = AssetType.objects.all()
    if asset_types.exists():
        asset_type_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(asset_types.values_list("name", flat=True))}"'
        )
        asset_type_validation.add(f'B2:B{len(example) + 1}')
        ws.add_data_validation(asset_type_validation)
    
    # Add data validation for company column
    companies = OwnerCompany.objects.all()
    if companies.exists():
        company_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(companies.values_list("name", flat=True))}"'
        )
        company_validation.add(f'F2:F{len(example) + 1}')
        ws.add_data_validation(company_validation)
    
    # Add data validation for status column
    status_validation = DataValidation(
        type="list",
        formula1=f'"{",".join([status[0] for status in ITAsset.STATUS_CHOICES])}"'
    )
    status_validation.add(f'N2:N{len(example) + 1}')
    ws.add_data_validation(status_validation)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=asset_template.xlsx'
    
    wb.save(response)
    return response

@login_required
def asset_upload(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            try:
                wb = openpyxl.load_workbook(excel_file)
                ws = wb.active
                
                # Skip header row
                for row in ws.iter_rows(min_row=2):
                    try:
                        # Check required fields
                        required_fields = {
                            'name': row[0].value,
                            'asset_type': row[1].value,
                            'serial_number': row[2].value,
                            'model': row[3].value,
                            'owner_company': row[5].value,
                            'mac_address_ethernet': row[8].value,
                            'status': row[13].value  # Added status field
                        }
                        
                        # Validate required fields
                        missing_fields = [field for field, value in required_fields.items() if not value]
                        if missing_fields:
                            messages.error(request, f'Row {row[0].row}: Missing required fields: {", ".join(missing_fields)}')
                            continue

                        # Get or create asset type
                        asset_type_name = row[1].value
                        try:
                            asset_type = AssetType.objects.get(name__iexact=asset_type_name)
                        except AssetType.DoesNotExist:
                            messages.error(request, f'Row {row[0].row}: Asset Type "{asset_type_name}" not found. Please create it first.')
                            continue

                        # Get the owner company
                        owner_company_name = row[5].value
                        try:
                            owner_company = OwnerCompany.objects.get(name__iexact=owner_company_name)
                        except OwnerCompany.DoesNotExist:
                            messages.error(request, f'Row {row[0].row}: Owner Company "{owner_company_name}" not found. Please create it first.')
                            continue

                        # Validate status
                        status = row[13].value
                        if status not in dict(ITAsset.STATUS_CHOICES):
                            messages.error(request, f'Row {row[0].row}: Invalid status "{status}". Please use one of the available statuses.')
                            continue

                        # Get the assigned employee if provided
                        assigned_employee_id = row[6].value
                        assigned_employee = None
                        if assigned_employee_id:
                            try:
                                assigned_employee = Employee.objects.get(employee_id=assigned_employee_id)
                                # Verify employee belongs to the owner company
                                if assigned_employee.company != owner_company:
                                    messages.error(request, f'Row {row[0].row}: Employee {assigned_employee_id} does not belong to {owner_company_name}. Please assign an employee from the correct company.')
                                    continue
                            except Employee.DoesNotExist:
                                messages.error(request, f'Row {row[0].row}: Employee with ID {assigned_employee_id} not found. Please use a valid employee ID.')
                                continue
                        
                        # Create asset
                        asset = ITAsset(
                            name=row[0].value,
                            asset_type=asset_type,
                            serial_number=row[2].value,
                            model=row[3].value,
                            manufacturer=row[4].value,
                            owner=owner_company,
                            assigned_to=assigned_employee,
                            mac_address_wifi=row[7].value,
                            mac_address_ethernet=row[8].value,
                            delivery_letter_code=row[9].value,
                            status=status,
                            purchase_date=row[10].value,
                            receipt_date=row[11].value,
                            warranty_expiry=row[12].value
                        )
                        asset.save()

                        # Create history record for new asset
                        AssetHistory.objects.create(
                            asset=asset,
                            action='received',
                            employee=assigned_employee,
                            notes=f'Asset received and added to inventory via bulk upload',
                            created_by=request.user
                        )

                        # If asset is assigned, create assignment history
                        if assigned_employee:
                            AssetHistory.objects.create(
                                asset=asset,
                                action='assigned',
                                employee=assigned_employee,
                                notes=f'Asset assigned to {assigned_employee.get_full_name()} during bulk upload',
                                created_by=request.user
                            )

                        messages.success(request, f'Row {row[0].row}: Asset "{asset.name}" created successfully.')
                    except Exception as e:
                        messages.error(request, f'Row {row[0].row}: Error processing row: {str(e)}')
                        continue
                
                return redirect('inventory:asset_list')
            except Exception as e:
                messages.error(request, f'Error uploading file: {str(e)}')
    
    return render(request, 'assets/asset_upload.html')

@login_required
def download_asset_data(request):
    wb = Workbook()
    ws = wb.active
    
    # Add headers
    headers = [
        'Asset Name',
        'Asset Type',
        'Serial Number',
        'Model',
        'Manufacturer',
        'Owner',
        'Wi-Fi MAC Address',
        'Ethernet MAC Address',
        'Delivery Letter Code',
        'Status',
        'Purchase Date',
        'Receipt Date',
        'Warranty Expiry'
    ]
    ws.append(headers)
    
    # Get filtered queryset using the same filters as the list view
    assets = ITAsset.objects.all()
    
    # Apply search filter
    search_query = request.GET.get('search', '')
    if search_query:
        assets = assets.filter(
            Q(name__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(model__icontains=search_query) |
            Q(delivery_letter_code__icontains=search_query)
        )
    
    # Apply other filters
    asset_type = request.GET.get('asset_type', '')
    if asset_type:
        assets = assets.filter(asset_type_id=asset_type)
    
    status = request.GET.get('status', '')
    if status:
        assets = assets.filter(status=status)
    
    manufacturer = request.GET.get('manufacturer', '')
    if manufacturer:
        assets = assets.filter(manufacturer=manufacturer)
    
    # Add asset data
    for asset in assets:
        row = [
            asset.name,
            asset.asset_type.name,  # Use the internal name for consistency
            asset.serial_number,
            asset.model,
            asset.manufacturer,
            f"{asset.assigned_to.get_full_name()} ({asset.assigned_to.company})" if asset.assigned_to else '',
            asset.mac_address_wifi,
            asset.mac_address_ethernet,
            asset.delivery_letter_code,
            asset.status,  # Use the internal status value for consistency
            asset.purchase_date,
            asset.receipt_date,
            asset.warranty_expiry
        ]
        ws.append(row)
    
    # Style the header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=assets.xlsx'
    
    wb.save(response)
    return response

# Asset Type Views
class AssetTypeListView(LoginRequiredMixin, ListView):
    model = AssetType
    template_name = 'assets/asset_type_list.html'
    context_object_name = 'asset_types'

class AssetTypeCreateView(LoginRequiredMixin, CreateView):
    model = AssetType
    template_name = 'assets/asset_type_form.html'
    fields = ['name', 'display_name']
    success_url = reverse_lazy('inventory:asset_type_list')

    def form_valid(self, form):
        messages.success(self.request, 'Asset Type created successfully.')
        return super().form_valid(form)

class AssetTypeUpdateView(LoginRequiredMixin, UpdateView):
    model = AssetType
    template_name = 'assets/asset_type_form.html'
    fields = ['name', 'display_name']
    success_url = reverse_lazy('inventory:asset_type_list')

    def form_valid(self, form):
        messages.success(self.request, 'Asset Type updated successfully.')
        return super().form_valid(form)

class AssetTypeDeleteView(LoginRequiredMixin, DeleteView):
    model = AssetType
    template_name = 'assets/asset_type_confirm_delete.html'
    success_url = reverse_lazy('inventory:asset_type_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Asset Type deleted successfully.')
        return super().delete(request, *args, **kwargs)

# Owner Company Views
class OwnerCompanyListView(LoginRequiredMixin, ListView):
    model = OwnerCompany
    template_name = 'company/owner_company_list.html'
    context_object_name = 'companies'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Owner Companies')
        return context

class OwnerCompanyCreateView(LoginRequiredMixin, CreateView):
    model = OwnerCompany
    form_class = OwnerCompanyForm
    template_name = 'company/owner_company_form.html'
    success_url = reverse_lazy('inventory:owner_company_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Add Company')
        return context

    def form_valid(self, form):
        messages.success(self.request, _('Owner Company created successfully.'))
        return super().form_valid(form)

class OwnerCompanyUpdateView(LoginRequiredMixin, UpdateView):
    model = OwnerCompany
    form_class = OwnerCompanyForm
    template_name = 'company/owner_company_form.html'
    success_url = reverse_lazy('inventory:owner_company_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Edit Company')
        return context

    def form_valid(self, form):
        messages.success(self.request, _('Owner Company updated successfully.'))
        return super().form_valid(form)

class OwnerCompanyDeleteView(LoginRequiredMixin, DeleteView):
    model = OwnerCompany
    template_name = 'company/owner_company_confirm_delete.html'
    success_url = reverse_lazy('inventory:owner_company_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Delete Company')
        return context

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Owner Company deleted successfully.'))
        return super().delete(request, *args, **kwargs)

class OutletListView(LoginRequiredMixin, ListView):
    model = Outlet
    template_name = 'outlets/outlet_list.html'
    context_object_name = 'outlets'
    paginate_by = 10

    def get_queryset(self):
        queryset = Outlet.objects.select_related('company', 'manager').all()
        
        # Get search and filter parameters
        search_query = self.request.GET.get('search', '')
        company_id = self.request.GET.get('company', '')
        status = self.request.GET.get('status', '')
        
        # Apply search filter if provided
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(address__icontains=search_query) |
                Q(manager__first_name__icontains=search_query) |
                Q(manager__last_name__icontains=search_query)
            )
        
        # Apply company filter if provided
        if company_id:
            queryset = queryset.filter(company_id=company_id)
        
        # Apply status filter if provided
        if status:
            is_active = status == 'active'
            queryset = queryset.filter(is_active=is_active)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Outlets')
        context['companies'] = OwnerCompany.objects.all()
        
        # Get selected company for filter badges
        company_id = self.request.GET.get('company', '')
        if company_id:
            context['selected_company'] = OwnerCompany.objects.filter(id=company_id).first()
        
        return context

class OutletDetailView(LoginRequiredMixin, DetailView):
    model = Outlet
    template_name = 'outlets/outlet_detail.html'
    context_object_name = 'outlet'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Outlet Details')
        return context

class OutletCreateView(LoginRequiredMixin, CreateView):
    model = Outlet
    template_name = 'outlets/outlet_form.html'
    fields = ['name', 'company', 'manager', 'address', 'google_maps_link', 'phone_number', 'email', 'is_active']

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Only show active employees as potential managers
        form.fields['manager'].queryset = Employee.objects.filter(is_active=True)
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Add Outlet')
        return context

    def form_valid(self, form):
        messages.success(self.request, _('Outlet created successfully.'))
        return super().form_valid(form)

class OutletUpdateView(LoginRequiredMixin, UpdateView):
    model = Outlet
    template_name = 'outlets/outlet_form.html'
    fields = ['name', 'company', 'manager', 'address', 'google_maps_link', 'phone_number', 'email', 'is_active']

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Only show active employees as potential managers
        form.fields['manager'].queryset = Employee.objects.filter(is_active=True)
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Edit Outlet')
        return context

    def form_valid(self, form):
        messages.success(self.request, _('Outlet updated successfully.'))
        return super().form_valid(form)

class OutletDeleteView(LoginRequiredMixin, DeleteView):
    model = Outlet
    template_name = 'outlets/outlet_confirm_delete.html'
    success_url = reverse_lazy('inventory:outlet_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Delete Outlet')
        return context

    def delete(self, request, *args, **kwargs):
        messages.success(request, _('Outlet deleted successfully.'))
        return super().delete(request, *args, **kwargs)

class EmployeePDFView(LoginRequiredMixin, DetailView):
    model = Employee
    template_name = 'employees/employee_pdf.html'
    
    def get(self, request, *args, **kwargs):
        employee = self.get_object()
        html_string = render_to_string(self.template_name, {'employee': employee})
        
        # Create PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="employee_{employee.employee_id}.pdf"'
        
        # Generate PDF
        pisa_status = pisa.CreatePDF(
            html_string,
            dest=response,
            encoding='utf-8'
        )
        
        if pisa_status.err:
            return HttpResponse('Error generating PDF')
        
        return response

class EmployeeListView(LoginRequiredMixin, ListView):
    model = Employee
    template_name = 'employees/employee_list.html'
    context_object_name = 'employees'
    paginate_by = 10

    def get_queryset(self):
        queryset = Employee.objects.select_related('department', 'company').all()
        
        # Get search and filter parameters
        search_query = self.request.GET.get('search', '')
        department_id = self.request.GET.get('department', '')
        company_id = self.request.GET.get('company', '')
        status = self.request.GET.get('status', '')
        
        # Apply search filter if provided
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(employee_id__icontains=search_query) |
                Q(national_id__icontains=search_query) |  # Add national_id to search
                Q(position__icontains=search_query)  # Add position to search
            )
        
        # Apply department filter if provided
        if department_id:
            queryset = queryset.filter(department_id=department_id)
        
        # Apply company filter if provided
        if company_id:
            queryset = queryset.filter(company_id=company_id)
        
        # Apply status filter if provided
        if status:
            is_active = status == 'active'
            queryset = queryset.filter(is_active=is_active)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add filter options
        context['departments'] = Department.objects.all()
        context['companies'] = OwnerCompany.objects.all()
        
        # Get selected department and company for filter badges
        department_id = self.request.GET.get('department', '')
        company_id = self.request.GET.get('company', '')
        
        if department_id:
            context['selected_department'] = Department.objects.filter(id=department_id).first()
        
        if company_id:
            context['selected_company'] = OwnerCompany.objects.filter(id=company_id).first()
        
        # Add pagination context
        paginator = context['paginator']
        page_obj = context['page_obj']
        
        # Get the current page number
        current_page = page_obj.number
        total_pages = paginator.num_pages
        
        # Calculate the range of page numbers to show
        if total_pages <= 5:
            page_range = range(1, total_pages + 1)
        else:
            if current_page <= 3:
                page_range = range(1, 6)
            elif current_page >= total_pages - 2:
                page_range = range(total_pages - 4, total_pages + 1)
            else:
                page_range = range(current_page - 2, current_page + 3)
        
        context.update({
            'is_paginated': page_obj.has_other_pages(),
            'page_obj': page_obj,
            'page_range': page_range,
            'current_page': current_page,
            'total_pages': total_pages,
        })
        
        return context

    def get(self, request, *args, **kwargs):
        if request.GET.get('export') == 'excel':
            # Get the filtered queryset
            queryset = self.get_queryset()
            
            # Generate dynamic filename based on filters
            filename_parts = ['employees']
            
            # Add department to filename
            department_id = request.GET.get('department')
            if department_id:
                try:
                    department = Department.objects.get(id=department_id)
                    filename_parts.append(department.name.lower())
                except Department.DoesNotExist:
                    pass
            
            # Add company to filename
            company_id = request.GET.get('company')
            if company_id:
                try:
                    company = OwnerCompany.objects.get(id=company_id)
                    filename_parts.append(company.code.lower())
                except OwnerCompany.DoesNotExist:
                    pass
            
            # Add status to filename
            status = request.GET.get('status')
            if status:
                filename_parts.append(status)
            
            # Add timestamp
            filename_parts.append(timezone.now().strftime('%Y%m%d'))
            
            # Create workbook and select active sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Filtered Employees"
            
            # Write headers
            headers = [
                'Employee ID',
                'National ID',
                'First Name',
                'Last Name',
                'Email',
                'Department',
                'Position',
                'Company',
                'Hire Date',
                'Status'
            ]
            ws.append(headers)
            
            # Write data
            for employee in queryset:
                row = [
                    employee.employee_id,
                    employee.national_id,
                    employee.first_name,
                    employee.last_name,
                    employee.email,
                    employee.department.name if employee.department else '',
                    employee.position,
                    employee.company.name if employee.company else '',
                    employee.hire_date.strftime('%Y-%m-%d') if employee.hire_date else '',
                    'Active' if employee.is_active else 'Inactive'
                ]
                ws.append(row)
            
            # Create the response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Join filename parts with underscores and add extension
            filename = '_'.join(filename_parts) + '.xlsx'
            response['Content-Disposition'] = f'attachment; filename={filename}'
            
            # Save the workbook to the response
            wb.save(response)
            return response
            
        return super().get(request, *args, **kwargs)

class EmployeeDetailView(LoginRequiredMixin, DetailView):
    model = Employee
    template_name = 'employees/employee_detail.html'
    context_object_name = 'employee'

class EmployeeCreateView(LoginRequiredMixin, CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'employees/employee_form.html'
    success_url = reverse_lazy('inventory:employee_list')

    def form_valid(self, form):
        messages.success(self.request, 'Employee created successfully.')
        return super().form_valid(form)

class EmployeeUpdateView(LoginRequiredMixin, UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'employees/employee_form.html'
    success_url = reverse_lazy('inventory:employee_list')

    def form_valid(self, form):
        messages.success(self.request, 'Employee updated successfully.')
        return super().form_valid(form)

class EmployeeDeleteView(LoginRequiredMixin, DeleteView):
    model = Employee
    template_name = 'employees/employee_confirm_delete.html'
    success_url = reverse_lazy('inventory:employee_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Employee deleted successfully.')
        return super().delete(request, *args, **kwargs)

@login_required
def owner_company_list(request):
    companies = OwnerCompany.objects.all()
    return render(request, 'company/owner_company_list.html', {
        'companies': companies,
        'title': 'Owner Companies'
    })

@login_required
def owner_company_detail(request, pk):
    company = get_object_or_404(OwnerCompany, pk=pk)
    return render(request, 'inventory/owner_company_detail.html', {
        'company': company,
        'title': f'Company: {company.name}'
    })

@login_required
def asset_type_list(request):
    asset_types = AssetType.objects.all()
    return render(request, 'inventory/asset_type_list.html', {
        'asset_types': asset_types,
        'title': 'Asset Types'
    })

@login_required
def asset_type_detail(request, pk):
    asset_type = get_object_or_404(AssetType, pk=pk)
    return render(request, 'inventory/asset_type_detail.html', {
        'asset_type': asset_type,
        'title': f'Asset Type: {asset_type.display_name}'
    })

# Department Views
@login_required
def department_list(request):
    departments = Department.objects.all()
    return render(request, 'department/department_list.html', {
        'departments': departments,
        'title': 'Departments'
    })

@login_required
def department_detail(request, pk):
    department = get_object_or_404(Department, pk=pk)
    return render(request, 'department/department_detail.html', {
        'department': department,
        'title': f'Department: {department.name}'
    })

@login_required
def department_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            department = Department.objects.create(name=name)
            messages.success(request, f'Department "{department.name}" created successfully.')
            return redirect('inventory:department_list')
    return render(request, 'department/department_form.html', {
        'title': 'Create Department'
    })

@login_required
def department_update(request, pk):
    department = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            department.name = name
            department.save()
            messages.success(request, f'Department "{department.name}" updated successfully.')
            return redirect('inventory:department_list')
    return render(request, 'department/department_form.html', {
        'department': department,
        'title': f'Update Department: {department.name}'
    })

@login_required
def department_delete(request, pk):
    department = get_object_or_404(Department, pk=pk)
    if request.method == 'POST':
        try:
            name = department.name
            department.delete()
            messages.success(request, f'Department "{name}" deleted successfully.')
            return redirect('inventory:department_list')
        except models.ProtectedError as e:
            employees = department.employees.all()
            positions = department.positions.all()
            messages.error(request, 'Cannot delete this department because it has associated employees and positions.')
            return render(request, 'department/department_confirm_delete.html', {
                'department': department,
                'title': f'Delete Department: {department.name}',
                'employees': employees,
                'positions': positions,
                'error': True
            })
    return render(request, 'department/department_confirm_delete.html', {
        'department': department,
        'title': f'Delete Department: {department.name}',
        'employees': department.employees.all(),
        'positions': department.positions.all()
    })

@login_required
def device_history(request, pk):
    device = get_object_or_404(ITAsset, pk=pk)
    device_history = device.device_history.all()
    
    context = {
        'title': f'{device.name} - History',
        'device': device,
        'device_history': device_history,
        'now': timezone.now(),
    }
    return render(request, 'inventory/device_history.html', context)

@login_required
def asset_history(request, pk):
    asset = get_object_or_404(ITAsset, pk=pk)
    asset_history = asset.asset_history.all()
    
    context = {
        'title': f'{asset.name} - History',
        'asset': asset,
        'asset_history': asset_history,
        'now': timezone.now(),
    }
    return render(request, 'assets/asset_history.html', context)

class NotificationListView(LoginRequiredMixin, ListView):
    model = Notification
    template_name = 'notification/notification_list.html'
    context_object_name = 'notifications'
    paginate_by = 15

    def get_queryset(self):
        queryset = Notification.objects.select_related(
            'category', 'employee_profile', 'content_type'
        )

        # Check if user has an employee profile
        if hasattr(self.request.user, 'employee_profile') and self.request.user.employee_profile:
            queryset = queryset.filter(
                Q(employee_profile=self.request.user.employee_profile) |
                Q(employee_profile__isnull=True)
            )
        else:
            # If user doesn't have an employee profile, only show notifications without a specific employee
            queryset = queryset.filter(employee_profile__isnull=True)

        # Filter by status
        status = self.request.GET.get('status')
        if status in ['unread', 'read', 'archived']:
            queryset = queryset.filter(status=status)
        elif not status:  # Default view excludes archived
            queryset = queryset.exclude(status='archived')

        # Filter by priority
        priority = self.request.GET.get('priority')
        if priority in ['low', 'medium', 'high', 'urgent']:
            queryset = queryset.filter(priority=priority)

        # Filter by category
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)

        # Filter by date range
        date_range = self.request.GET.get('date_range')
        if date_range:
            today = timezone.now().date()
            if date_range == 'today':
                queryset = queryset.filter(created_at__date=today)
            elif date_range == 'week':
                queryset = queryset.filter(created_at__date__gte=today - timezone.timedelta(days=7))
            elif date_range == 'month':
                queryset = queryset.filter(created_at__date__gte=today - timezone.timedelta(days=30))

        # Search
        search_query = self.request.GET.get('q')
        if search_query:
            queryset = queryset.filter(
                Q(title__icontains=search_query) |
                Q(message__icontains=search_query)
            )

        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get notification statistics
        notifications = self.get_queryset()
        context['total_count'] = notifications.count()
        context['unread_count'] = notifications.filter(status='unread').count()
        context['archived_count'] = notifications.filter(status='archived').count()
        
        # Get counts by priority
        priority_counts = notifications.values('priority').annotate(
            count=Count('id')
        ).order_by('priority')
        context['priority_counts'] = {
            item['priority']: item['count'] for item in priority_counts
        }
        
        # Get counts by category
        category_counts = notifications.values(
            'category__name', 'category__icon', 'category__color'
        ).annotate(
            count=Count('id')
        ).order_by('category__name')
        context['category_counts'] = category_counts
        
        # Add filter values to context
        context['current_status'] = self.request.GET.get('status', '')
        context['current_priority'] = self.request.GET.get('priority', '')
        context['current_category'] = self.request.GET.get('category', '')
        context['current_date_range'] = self.request.GET.get('date_range', '')
        context['search_query'] = self.request.GET.get('q', '')
        
        # Add categories for filter dropdown
        context['categories'] = NotificationCategory.objects.all()
        
        return context

class NotificationDetailView(LoginRequiredMixin, DetailView):
    model = Notification
    template_name = 'notification/notification_detail.html'
    context_object_name = 'notification'

    def get_queryset(self):
        queryset = Notification.objects.all()
        
        # Check if user has an employee profile
        if hasattr(self.request.user, 'employee_profile') and self.request.user.employee_profile:
            queryset = queryset.filter(
                Q(employee_profile=self.request.user.employee_profile) |
                Q(employee_profile__isnull=True)
            )
        else:
            # If user doesn't have an employee profile, only show notifications without a specific employee
            queryset = queryset.filter(employee_profile__isnull=True)
        
        return queryset

    def get(self, request, *args, **kwargs):
        response = super().get(request, *args, **kwargs)
        # Mark as read when viewed
        self.object.mark_as_read()
        return response

@login_required
@require_POST
def mark_notification_read(request, pk):
    # Build the filter based on user's employee profile
    if hasattr(request.user, 'employee_profile') and request.user.employee_profile:
        notification_filter = Q(
            Q(employee_profile=request.user.employee_profile) |
            Q(employee_profile__isnull=True)
        )
    else:
        notification_filter = Q(employee_profile__isnull=True)
    
    notification = get_object_or_404(
        Notification,
        notification_filter,
        pk=pk
    )
    notification.mark_as_read()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'success'})
    
    messages.success(request, _('Notification marked as read.'))
    return redirect('inventory:notification_list')

@login_required
@require_POST
def mark_notification_unread(request, pk):
    # Build the filter based on user's employee profile
    if hasattr(request.user, 'employee_profile') and request.user.employee_profile:
        notification_filter = Q(
            Q(employee_profile=request.user.employee_profile) |
            Q(employee_profile__isnull=True)
        )
    else:
        notification_filter = Q(employee_profile__isnull=True)
    
    notification = get_object_or_404(
        Notification,
        notification_filter,
        pk=pk
    )
    notification.mark_as_unread()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'success'})
    
    messages.success(request, _('Notification marked as unread.'))
    return redirect('inventory:notification_list')

@login_required
@require_POST
def archive_notification(request, pk):
    # Build the filter based on user's employee profile
    if hasattr(request.user, 'employee_profile') and request.user.employee_profile:
        notification_filter = Q(
            Q(employee_profile=request.user.employee_profile) |
            Q(employee_profile__isnull=True)
        )
    else:
        notification_filter = Q(employee_profile__isnull=True)
    
    notification = get_object_or_404(
        Notification,
        notification_filter,
        pk=pk
    )
    notification.archive()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'success'})
    
    messages.success(request, _('Notification archived.'))
    return redirect('inventory:notification_list')

@login_required
@require_POST
def unarchive_notification(request, pk):
    # Build the filter based on user's employee profile
    if hasattr(request.user, 'employee_profile') and request.user.employee_profile:
        notification_filter = Q(
            Q(employee_profile=request.user.employee_profile) |
            Q(employee_profile__isnull=True)
        )
    else:
        notification_filter = Q(employee_profile__isnull=True)
    
    notification = get_object_or_404(
        Notification,
        notification_filter,
        pk=pk
    )
    notification.unarchive()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'success'})
    
    messages.success(request, _('Notification unarchived.'))
    return redirect('inventory:notification_list')

@login_required
@require_POST
def mark_all_read(request):
    # Build the filter based on user's employee profile
    if hasattr(request.user, 'employee_profile') and request.user.employee_profile:
        notification_filter = Q(
            Q(employee_profile=request.user.employee_profile) |
            Q(employee_profile__isnull=True)
        )
    else:
        notification_filter = Q(employee_profile__isnull=True)
    
    Notification.objects.filter(
        notification_filter,
        status='unread'
    ).update(
        status='read',
        read_at=timezone.now()
    )
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'success'})
    
    messages.success(request, _('All notifications marked as read.'))
    return redirect('inventory:notification_list')

@login_required
def get_notifications_ajax(request):
    """AJAX endpoint for real-time notification updates"""
    # Build the filter based on user's employee profile
    if hasattr(request.user, 'employee_profile') and request.user.employee_profile:
        notification_filter = Q(
            Q(employee_profile=request.user.employee_profile) |
            Q(employee_profile__isnull=True)
        )
    else:
        notification_filter = Q(employee_profile__isnull=True)
    
    # Get unread count first
    unread_count = Notification.objects.filter(
        notification_filter,
        status='unread'
    ).count()
    
    # Get both read and unread notifications, but prioritize unread ones
    notifications = Notification.objects.filter(
        notification_filter
    ).exclude(
        status='archived'
    ).select_related(
        'category'
    ).order_by(
        '-status',  # This will put unread first since 'unread' > 'read'
        '-created_at'
    )[:10]  # Limit to 10 most recent notifications
    
    context = {
        'notifications': notifications,
        'unread_notifications_count': unread_count
    }
    
    html = render_to_string(
        'inventory/includes/notification_dropdown.html',
        context,
        request=request
    )
    
    return JsonResponse({
        'html': html,
        'unread_count': unread_count
    })

class UserPermissionsView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'inventory/user_permissions.html'
    
    def test_func(self):
        return self.request.user.is_superuser or self.request.user.groups.filter(name='Admin').exists()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['users'] = User.objects.select_related('employee_profile').prefetch_related('groups').all()
        context['groups'] = Group.objects.all()
        context['available_permissions'] = Permission.objects.select_related('content_type').all()
        return context

    def post(self, request, *args, **kwargs):
        user_id = request.POST.get('user_id')
        action = request.POST.get('action')
        
        if not user_id or not action:
            messages.error(request, _('Invalid request parameters'))
            return redirect('inventory:user_permissions')
            
        try:
            user = User.objects.get(id=user_id)
            
            if action == 'update_groups':
                group_ids = request.POST.getlist('groups')
                user.groups.set(Group.objects.filter(id__in=group_ids))
                messages.success(request, _(f'Groups updated for {user.username}'))
                
            elif action == 'update_permissions':
                permission_ids = request.POST.getlist('permissions')
                user.user_permissions.set(Permission.objects.filter(id__in=permission_ids))
                messages.success(request, _(f'Permissions updated for {user.username}'))
                
            elif action == 'toggle_active':
                user.is_active = not user.is_active
                user.save()
                status = 'activated' if user.is_active else 'deactivated'
                messages.success(request, _(f'User {user.username} has been {status}'))
                
            elif action == 'toggle_staff':
                if request.user.is_superuser:  # Only superusers can change staff status
                    user.is_staff = not user.is_staff
                    user.save()
                    status = 'granted' if user.is_staff else 'revoked'
                    messages.success(request, _(f'Staff status {status} for {user.username}'))
                else:
                    messages.error(request, _('You do not have permission to change staff status'))
                    
        except User.DoesNotExist:
            messages.error(request, _('User not found'))
        except Exception as e:
            messages.error(request, str(e))
            
        return redirect('inventory:user_permissions')

@login_required
def global_asset_history(request):
    history = AssetHistory.objects.select_related('asset', 'employee', 'created_by').order_by('-date')
    return render(request, 'reports/global_asset_history.html', {
        'history': history,
        'title': _('Global Asset History')
    })

@login_required
def reports_dashboard(request):
    """View for displaying comprehensive system reports and analytics."""
    # Get filter parameters
    report_type = request.GET.get('type', 'full')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    try:
        if date_from:
            date_from = datetime.strptime(date_from, '%Y-%m-%d')
        if date_to:
            date_to = datetime.strptime(date_to, '%Y-%m-%d')
            # Include the entire day
            date_to = date_to + timedelta(days=1)
    except ValueError:
        messages.error(request, _('Invalid date format'))
        return redirect('inventory:reports_dashboard')

    # Asset statistics
    total_assets = ITAsset.objects.count()
    available_assets = ITAsset.objects.filter(status='available').count()
    assigned_assets = ITAsset.objects.filter(status='assigned').count()
    maintenance_assets = ITAsset.objects.filter(status='maintenance').count()
    retired_assets = ITAsset.objects.filter(status='retired').count()
    
    # Department statistics
    departments = Department.objects.annotate(
        asset_count=Count('employees__itasset', distinct=True)
    )
    
    # Device Model statistics
    model_stats = ITAsset.objects.values('model', 'asset_type__display_name') \
        .annotate(count=Count('id')) \
        .filter(count__gt=0) \
        .order_by('-count')
    
    # Recent activity
    recent_activities = AssetHistory.objects.select_related(
        'asset', 'employee'
    ).order_by('-date')[:10]

    # Calculate YoY growth
    today = timezone.now()
    one_year_ago = today - timedelta(days=365)
    assets_one_year_ago = ITAsset.objects.filter(purchase_date__lte=one_year_ago).count()
    current_assets = ITAsset.objects.count()
    yoy_growth = ((current_assets - assets_one_year_ago) / assets_one_year_ago * 100) if assets_one_year_ago > 0 else 0

    # Calculate 3-month average growth rate
    three_months_ago = today - timedelta(days=90)
    monthly_growth = ITAsset.objects.filter(
        purchase_date__gte=three_months_ago
    ).annotate(
        month=TruncMonth('purchase_date')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')
    
    avg_growth_rate = sum(m['count'] for m in monthly_growth) / 3 if monthly_growth else 0

    # Company device statistics
    company_device_stats = []
    
    # Get all companies, even those without devices
    for company in OwnerCompany.objects.all().order_by('name'):
        # Get base queryset for this company
        company_assets = ITAsset.objects.filter(owner=company)
        total_devices = company_assets.count()
        
        # Always include the company, even if it has no devices
        company_data = {
            'company': company,
            'total': total_devices,
            'available': company_assets.filter(status='available').count(),
            'assigned': company_assets.filter(status='assigned').count(),
            'maintenance': company_assets.filter(status='maintenance').count(),
            'retired': company_assets.filter(status='retired').count(),
            'available_percentage': 0,
            'assigned_percentage': 0,
            'maintenance_percentage': 0,
            'retired_percentage': 0,
            'device_types': []
        }
        
        # Calculate percentages only if there are devices
        if total_devices > 0:
            company_data.update({
                'available_percentage': (company_data['available'] / total_devices * 100),
                'assigned_percentage': (company_data['assigned'] / total_devices * 100),
                'maintenance_percentage': (company_data['maintenance'] / total_devices * 100),
                'retired_percentage': (company_data['retired'] / total_devices * 100),
            })
            
            # Get device types distribution
            for asset_type in AssetType.objects.all():
                type_count = company_assets.filter(asset_type=asset_type).count()
                if type_count > 0:
                    type_percentage = (type_count / total_devices * 100)
                    company_data['device_types'].append({
                        'type': asset_type.display_name,
                        'count': type_count,
                        'percentage': type_percentage
                    })
            
            # Sort device types by count in descending order
            company_data['device_types'].sort(key=lambda x: x['count'], reverse=True)
        
        company_device_stats.append(company_data)
    
    # Calculate Asset Growth Analysis data
    twelve_months_ago = today - timedelta(days=365)
    
    # Monthly asset acquisition data
    monthly_acquisitions = ITAsset.objects.filter(
        purchase_date__gte=twelve_months_ago
    ).annotate(
        month=TruncMonth('purchase_date')
    ).values('month').annotate(
        new_assets=Count('id')
    ).order_by('month')
    
    # Calculate growth rates
    growth_data = []
    previous_count = ITAsset.objects.filter(purchase_date__lt=twelve_months_ago).count()
    
    for acquisition in monthly_acquisitions:
        current_count = previous_count + acquisition['new_assets']
        if previous_count > 0:
            growth_rate = ((current_count - previous_count) / previous_count) * 100
        else:
            growth_rate = 100 if current_count > 0 else 0
            
        growth_data.append({
            'month': acquisition['month'].strftime('%B %Y'),
            'new_assets': acquisition['new_assets'],
            'total_assets': current_count,
            'growth_rate': round(growth_rate, 2)
        })
        previous_count = current_count
    
    # Calculate trend indicators
    if len(growth_data) >= 2:
        last_month_growth = growth_data[-1]['growth_rate']
        prev_month_growth = growth_data[-2]['growth_rate']
        trend = 'up' if last_month_growth > prev_month_growth else 'down'
    else:
        trend = 'neutral'
    
    context = {
        'title': _('Reports Dashboard'),
        'report_type': report_type,
        'date_from': date_from.strftime('%Y-%m-%d') if date_from else '',
        'date_to': (date_to - timedelta(days=1)).strftime('%Y-%m-%d') if date_to else '',
        'report_types': [
            ('full', _('Full Report')),
            ('warranty', _('Warranty Report')),
            ('assignments', _('Assignments Report')),
        ],
        'total_assets': total_assets,
        'available_assets': available_assets,
        'assigned_assets': assigned_assets,
        'maintenance_assets': maintenance_assets,
        'retired_assets': retired_assets,
        'departments': departments,
        'model_stats': model_stats,
        'recent_activities': recent_activities,
        'yoy_growth': round(yoy_growth, 2),
        'avg_growth_rate': round(avg_growth_rate, 2),
        'company_device_stats': company_device_stats,
        'growth_data': growth_data,
        'growth_trend': trend,
        'total_growth_rate': growth_data[-1]['growth_rate'] if growth_data else 0,
    }
    
    return render(request, 'reports/reports_dashboard.html', context)

@login_required
def generate_report_pdf(request):
    """Generate a PDF report based on the specified parameters."""
    # Get report parameters
    report_type = request.GET.get('type', 'full')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    try:
        if date_from:
            date_from = datetime.strptime(date_from, '%Y-%m-%d')
        if date_to:
            date_to = datetime.strptime(date_to, '%Y-%m-%d')
            # Include the entire day
            date_to = date_to + timedelta(days=1)
    except ValueError:
        messages.error(request, _('Invalid date format'))
        return redirect('inventory:reports_dashboard')

    # Get the report data
    context = {
        'title': _('Asset Management Report'),
        'report_type': report_type,
        'date_from': date_from.strftime('%Y-%m-%d') if date_from else None,
        'date_to': (date_to - timedelta(days=1)).strftime('%Y-%m-%d') if date_to else None,
        'generated_at': timezone.now(),
        'generated_by': request.user.get_full_name() or request.user.username,
    }

    # Add statistics based on report type
    if report_type == 'full' or report_type == 'assignments':
        # Asset statistics
        context.update({
            'total_assets': ITAsset.objects.count(),
            'available_assets': ITAsset.objects.filter(status='available').count(),
            'assigned_assets': ITAsset.objects.filter(status='assigned').count(),
            'maintenance_assets': ITAsset.objects.filter(status='maintenance').count(),
            'retired_assets': ITAsset.objects.filter(status='retired').count(),
        })

        # Recent assignments
        recent_assignments = AssetHistory.objects.filter(
            action='assigned'
        ).select_related(
            'asset', 'employee'
        ).order_by('-date')[:20]
        context['recent_assignments'] = recent_assignments

    if report_type == 'full' or report_type == 'warranty':
        # Warranty expiry statistics
        today = timezone.now().date()
        thirty_days = today + timedelta(days=30)
        ninety_days = today + timedelta(days=90)

        context.update({
            'expired_warranty': ITAsset.objects.filter(
                warranty_expiry__lt=today
            ).exclude(status='retired').count(),
            'expiring_30_days': ITAsset.objects.filter(
                warranty_expiry__gte=today,
                warranty_expiry__lte=thirty_days
            ).exclude(status='retired').count(),
            'expiring_90_days': ITAsset.objects.filter(
                warranty_expiry__gt=thirty_days,
                warranty_expiry__lte=ninety_days
            ).exclude(status='retired').count(),
        })

        # Assets with expiring warranty
        expiring_assets = ITAsset.objects.filter(
            warranty_expiry__gte=today,
            warranty_expiry__lte=ninety_days
        ).exclude(
            status='retired'
        ).order_by('warranty_expiry')
        context['expiring_assets'] = expiring_assets

    # Company statistics
    companies = OwnerCompany.objects.all()
    company_stats = []
    for company in companies:
        assets = ITAsset.objects.filter(owner=company)
        company_stats.append({
            'name': company.name,
            'total_assets': assets.count(),
            'assigned_assets': assets.filter(status='assigned').count(),
            'available_assets': assets.filter(status='available').count(),
            'maintenance_assets': assets.filter(status='maintenance').count(),
            'retired_assets': assets.filter(status='retired').count(),
        })
    context['company_stats'] = company_stats

    # Render the HTML template
    template = get_template('reports/pdf_report.html')
    html = template.render(context)
    
    # Create a file-like buffer to receive PDF data
    buffer = BytesIO()
    
    # Convert HTML to PDF
    pisa_status = pisa.CreatePDF(html, dest=buffer)
    
    # If error creating PDF, return error response
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)
    
    # Get the value of the BytesIO buffer and write it to the response
    pdf = buffer.getvalue()
    buffer.close()
    
    # Create the HTTP response
    response = HttpResponse(content_type='application/pdf')
    filename = f'inventory_report_{report_type}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write(pdf)
    
    return response

class DatabaseBackupView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'database/backup.html'

    def test_func(self):
        return self.request.user.is_staff

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Database Backup')
        return context

class DatabaseRestoreView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'database/restore.html'

    def test_func(self):
        return self.request.user.is_staff

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Database Restore')
        return context

@login_required
def asset_receipt(request, pk):
    """View for displaying and handling the asset receipt form."""
    asset = get_object_or_404(ITAsset, pk=pk)
    
    # Only show receipt for assigned assets
    if asset.status != 'assigned' or not asset.assigned_to:
        messages.error(request, 'Receipt can only be generated for assigned assets.')
        return redirect('inventory:asset_detail', pk=pk)
    
    context = {
        'asset': asset,
        'company_name_ar': asset.owner.name_ar if hasattr(asset.owner, 'name_ar') else asset.owner.name,
        'admin_name': request.user.get_full_name() or request.user.username,
    }
    
    return render(request, 'assets/receipt_form.html', context)

@login_required
def asset_unassign(request, pk):
    if request.method == 'POST':
        asset = get_object_or_404(ITAsset, pk=pk)
        employee = asset.assigned_to
        if employee:
            # Store employee ID before unassigning
            employee_id = employee.id
            # Unassign the asset
            asset.assigned_to = None
            asset.status = 'available'  # Update status to available
            asset.save()
            
            # Add success message
            messages.success(request, _('Asset successfully unassigned.'))
            
            # Redirect back to the employee detail page
            return redirect('inventory:employee_detail', pk=employee_id)
    
    # If not POST or no employee, redirect to asset list
    return redirect('inventory:asset_list')

class SuperUserRegistrationView(CreateView):
    template_name = 'registration/register.html'
    form_class = SuperUserRegistrationForm
    success_url = reverse_lazy('login')

    def form_valid(self, form):
        user = form.save(commit=False)
        user.is_superuser = True
        user.is_staff = True
        user.save()
        messages.success(self.request, _('Your account has been created successfully. Please log in.'))
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Register')
        return context

class AboutView(TemplateView):
    template_name = 'index/about.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('About Us')
        return context

class DocsView(TemplateView):
    template_name = 'index/docs.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Documentation')
        return context

class PrivacyView(TemplateView):
    template_name = 'index/privacy.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Privacy Policy')
        return context

class TermsView(TemplateView):
    template_name = 'index/terms.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Terms of Service')
        return context

class EmployeeListContentView(EmployeeListView):
    template_name = 'employees/employee_list_content.html'

    def get(self, request, *args, **kwargs):
        # Only serve HTMX requests for the content template
        if not request.headers.get('HX-Request'):
            return HttpResponseBadRequest('This view only serves HTMX requests')
            
        self.object_list = self.get_queryset()
        context = self.get_context_data()
        return render(request, self.template_name, context)

class OutletToggleStatusView(LoginRequiredMixin, View):
    def post(self, request, pk):
        outlet = get_object_or_404(Outlet, pk=pk)
        outlet.is_active = not outlet.is_active
        outlet.save()
        messages.success(request, _('Outlet status updated successfully.'))
        return redirect('inventory:outlet_detail', pk=pk)

class AssetUnassignOutletView(LoginRequiredMixin, View):
    def post(self, request, pk):
        asset = get_object_or_404(ITAsset, pk=pk)
        outlet = asset.outlet
        if outlet:
            asset.outlet = None
            asset.save()
            messages.success(request, _('Asset unassigned from outlet successfully.'))
            return redirect('inventory:outlet_detail', pk=outlet.pk)
        messages.error(request, _('Asset is not assigned to any outlet.'))
        return redirect('inventory:asset_detail', pk=pk)

class AssetAssignView(LoginRequiredMixin, View):
    template_name = 'assets/asset_assign.html'

    def get(self, request):
        form = AssetAssignForm()
        outlet_id = request.GET.get('outlet')
        if outlet_id:
            outlet = get_object_or_404(Outlet, pk=outlet_id)
            form.fields['outlet'].initial = outlet.id
        employee_id = request.GET.get('employee')
        if employee_id:
            employee = get_object_or_404(Employee, pk=employee_id)
            form.fields['employee'].initial = employee.id
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = AssetAssignForm(request.POST)
        if form.is_valid():
            asset = form.cleaned_data['asset']
            employee = form.cleaned_data['employee']
            outlet = form.cleaned_data['outlet']
            
            # Update asset assignment
            asset.assigned_to = employee
            asset.outlet = outlet
            asset.status = 'assigned'
            asset.save()

            # Create asset history entry
            AssetHistory.objects.create(
                asset=asset,
                action='assigned',
                employee=employee,
                notes=f'Assigned to {employee.get_full_name() if employee else "N/A"}'
                      f'{" and " + outlet.name if outlet else ""}',
                created_by=request.user
            )

            messages.success(request, _('Asset assigned successfully.'))
            if outlet:
                return redirect('inventory:outlet_detail', pk=outlet.pk)
            elif employee:
                return redirect('inventory:employee_detail', pk=employee.pk)
            return redirect('inventory:asset_detail', pk=asset.pk)
        return render(request, self.template_name, {'form': form})

class AssetUnassignView(LoginRequiredMixin, View):
    def post(self, request, pk):
        asset = get_object_or_404(ITAsset, pk=pk)
        employee = asset.assigned_to
        if employee:
            asset.assigned_to = None
            asset.status = 'available'
            asset.save()
            messages.success(request, _('Asset unassigned successfully.'))
            return redirect('inventory:employee_detail', pk=employee.pk)
        messages.error(request, _('Asset is not assigned to any employee.'))
        return redirect('inventory:asset_detail', pk=pk)

class UserAccountView(LoginRequiredMixin, TemplateView):
    template_name = 'user_profile/user_account.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Account Settings')
        return context

class UserProfileView(LoginRequiredMixin, TemplateView):
    template_name = 'user_profile/profile.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('User Profile')
        return context

class LandingPageView(TemplateView):
    template_name = 'index/landing.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('IT Asset Management Solutions')
        return context

class ContactView(TemplateView):
    template_name = 'index/contact.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Contact Us')
        return context

@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important to keep user logged in
            messages.success(request, _('Your password was successfully updated!'))
            return redirect('inventory:user_account')
        else:
            messages.error(request, _('Please correct the error below.'))
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'user_profile/change_password.html', {
        'form': form,
        'title': _('Change Password')
    })

class RegisterView(CreateView):
    form_class = UserRegistrationForm
    template_name = 'registration/register.html'
    success_url = reverse_lazy('inventory:home')

    def form_valid(self, form):
        response = super().form_valid(form)
        login(self.request, self.object)  # Log in the user after registration
        messages.success(self.request, _('Account created successfully! Welcome to Inventory Management System.'))
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = _('Create Account')
        return context

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            messages.info(request, _('You are already logged in.'))
            return redirect('inventory:home')
        return super().dispatch(request, *args, **kwargs)